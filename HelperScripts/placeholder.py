
import sys
import re
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
from pathlib import Path
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows
import tempfile

def extract_group_data_from_txt(folder: Path):
    pattern = re.compile(r'([A-Z]{2})_(\d{2,3})_(bL|3L)_group_acc\.txt$', re.IGNORECASE)
    records = []
    for file in folder.rglob("*_group_acc.txt"):
        match = pattern.search(file.name)
        if not match:
            continue
        label, distance, lighting = match.groups()
        with open(file, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        mae = None
        matches = None
        for line in lines:
            if "Mean Angular Error" in line:
                mae = float(line.split(":")[1].strip())
            elif "Matched timestamps" in line:
                matches = int(line.split(":")[1].strip())
        if mae is not None and matches is not None:
            records.append({
                "Label": label,
                "Distance": int(distance),
                "Lighting": lighting,
                "MAE": mae,
                "Matches": matches
            })
    df = pd.DataFrame(records)
    return df

def create_pivot_heatmap(df, title, output_image_path):
    pivot = df.pivot_table(index="Label", columns=["Distance", "Lighting"], values="MAE")
    plt.figure(figsize=(12, 6))
    sns.heatmap(pivot, annot=True, fmt=".2f", cmap="coolwarm", cbar_kws={'label': 'MAE (°)'})
    plt.title(title)
    plt.tight_layout()
    plt.savefig(output_image_path)
    plt.close()
    return pivot

if __name__ == "__main__":
    if len(sys.argv) != 4:
        print("Usage: python export_group_mae_to_excel.py <group_dir_1> <group_dir_2> <output_excel_path>")
        sys.exit(1)

    dir1 = Path(sys.argv[1])
    dir2 = Path(sys.argv[2])
    output_path = Path(sys.argv[3])

    df1 = extract_group_data_from_txt(dir1)
    df2 = extract_group_data_from_txt(dir2)

    with tempfile.TemporaryDirectory() as tmpdir:
        tmp1 = Path(tmpdir) / "heatmap_521.png"
        tmp2 = Path(tmpdir) / "heatmap_581.png"

        pivot1 = create_pivot_heatmap(df1, "Camera 521 - Mean Angular Error", tmp1)
        pivot2 = create_pivot_heatmap(df2, "Camera 581 - Mean Angular Error", tmp2)

        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Camera_521_Heatmap"
        img1 = XLImage(tmp1)
        ws1.add_image(img1, "A1")

        ws2 = wb.create_sheet("Camera_581_Heatmap")
        img2 = XLImage(tmp2)
        ws2.add_image(img2, "A1")

        # Optionally include raw tables
        ws3 = wb.create_sheet("Camera_521_Data")
        for r in dataframe_to_rows(pivot1.reset_index(), index=False, header=True):
            ws3.append(r)

        ws4 = wb.create_sheet("Camera_581_Data")
        for r in dataframe_to_rows(pivot2.reset_index(), index=False, header=True):
            ws4.append(r)

        wb.save(output_path)
        print(f"Excel with heatmaps saved to {output_path}")
